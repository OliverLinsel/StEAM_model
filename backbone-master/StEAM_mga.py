#%%
import sys
import subprocess as sp
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from gams import GamsWorkspace, GamsDatabase
from backbonetools.io import BackboneResult
from backbonetools.io import BackboneInput
import shutil
import os

# original code written by KE and KT

# working_path = r'/mnt/speicher/.wissmit/oliver/Data/Backbone/steam_mga'
# define root path
working_path = r"C:\Users\oliver\Documents\RUB\01_Projekte\StEAM\Programme\StEAM_model\backbone-master"
sys.path.append(working_path)

#create input and output folder paths
input_folder_path = f"{working_path}/input/mga_input"
output_folder_path = f"{working_path}/output/mga_output"

# # Construct the path to Backbone.gms using os.path.join
# backbone_gms_path = os.path.join(working_path, "Backbone.gms")
# print(f"Path to Backbone.gms: {backbone_gms_path}")  # Print the path for verification

# # Verify that the file exists
# if not os.path.exists(backbone_gms_path):
#     raise FileNotFoundError(f"The file {backbone_gms_path} does not exist.")
# gams_executable = r"C:\GAMS\51\gams.exe"

#### copying all necessary optimization files to the mga_input folder structure to isolate from the normal system runs ####

# Ensure the target directories exist
os.makedirs(input_folder_path, exist_ok=True)
os.makedirs(output_folder_path, exist_ok=True)

# List of files to copy
files_to_copy = ['cplex.opt', 'gurobi.opt', 'investInit.gms', 'timeAndSamples.inc', "modelsInit.gms"]

# Copy each file to mga_input
for file in files_to_copy:
    source_file = os.path.join(working_path, "input", file)
    target_file = os.path.join(input_folder_path, file)
    shutil.copy2(source_file, target_file)
    print(f"Copied {file} to {input_folder_path}")

# Copy inputData.gdx as inputData_MGA.gdx to mga_input
source_gdx = os.path.join(working_path, "input", 'inputData.gdx')
target_gdx = os.path.join(input_folder_path, 'inputData_MGA.gdx')
shutil.copy2(source_gdx, target_gdx)
print(f"Copied inputData.gdx as inputData_MGA.gdx to {input_folder_path}")

#%%
#### definition of functions ####

# run Backbone with different input directories (one for cost obj. and one for generation obj. with cost constraint)
def runBackbone(input_dir, input_file, output_dir, output_file, costConstraint=False, maxCost="inf"):
    if costConstraint == False:
        result = sp.run(["gams", str(working_path) + "\Backbone.gms", r"--debug=1", r"--input_dir={}".format(input_dir), r"--input_file_gdx={}".format(input_file), r"--output_dir={}".format(output_dir), r"--output_file={}".format(output_file)], stdout=sp.PIPE)
        return result
    elif costConstraint == True:
        result = sp.run(["gams", str(working_path) + "\Backbone.gms", r"--debug=1", r"--input_dir={}".format(input_dir), r"--input_file_gdx={}".format(input_file), r"--output_dir={}".format(output_dir), r"--output_file={}".format(output_file), r"--maxTotalCost={}".format(maxCost)], stdout=sp.PIPE)
        return result

# # read cost from Backbone result
# def readCostResult(output_dir, file_name): 
#     bb_result = BackboneResult(r"{}\{}".format(output_dir, file_name)) # import backbone data 
#     cost = round(bb_result.param_as_df("r_totalRealizedCost")["value"], 0) # read parameter from results file
#     return cost

# # read and calculate total generation share (of total demand) per technology from Backbone result
# def bbResultToTableNodeTechnology(resultFile, paramName):

#     df = resultFile.param_as_df(paramName)                              # read result
#     if "grid" in df.columns:
#         df = df[df.grid == 'elec']                                      # elec grid only

#     dfUnit = df["unit"].str.split(" ", expand=True)                     # create new df with splitted unit name
#     for c in dfUnit.index:   
#         if dfUnit.columns.stop > 3 and dfUnit.loc[c, 3] != None:
#             if "invest" in dfUnit.loc[c, 3]:
#                 dfUnit.at[c, 2] = dfUnit.loc[c, 2] + dfUnit.loc[c, 3]
#     df["technology"] = dfUnit[2]                                        # write column technology to df
#     if "node" not in df.columns:
#         df["node"] = dfUnit[0] + " " + dfUnit[1]                        # extract node name from unit name

#     if "grid" in df.columns: df.drop(columns="grid", inplace=True)      # drop columns "grid" and "unit"
#     if "unit" in df.columns: df.drop(columns="unit", inplace=True)

#     df = df.pivot_table(index="node", columns="technology", values="Val")   # create table "node" x "technology"

#     df[abs(df) < 1.e-6] = 0.                                            # replace absolute values < 1e-6 with 0
#     df.fillna(0., inplace=True)                                         # replace NaNs with 0
    
#     return df

# # write weights dataframe to input data and save with new name
# def writeWeightsToInputData(input_dir, inputFile, output_file):  # do not parallelise excel stuff, doesn´t really matter for runtime and might cause trouble as read multiple times
#     input_data = load_workbook(filename = r"{}\{}".format(input_dir, inputFile))   # load input data .xlsx
#     p_groupPolicy = input_data['p_groupPolicy']
#     for i in range(0, len(groups.index)):
#         for c in ["A", "B", "C"]: 
#             if c == "A":
#                 p_groupPolicy['A{}'.format(i+2)] = weights.iloc[i,0]
#             elif c == "B":
#                 p_groupPolicy['B{}'.format(i+2)] = weights.iloc[i,1]
#             elif c == "C":
#                 p_groupPolicy['C{}'.format(i+2)] = weights.iloc[i,2]
#     # input_data.save("{}\{}".format(input_dir, output_file))                        # save file under name containing emission cap
#     input_data.save("{}/{}".format(input_dir, output_file))  # Use forward slashes

def writeWeightsToInputDataInBetter(input_dir, input_file, output_file, my_weights, my_groups): #less of a criminal act, since we dont use global variables...
    input_data_file = BackboneInput(os.path.join(input_dir, input_file))
    # input_data = load_workbook(filename = f"{input_dir}{input_file}")   # load input data .xlsx
    p_groupPolicy = input_data_file.p_groupPolicy() #param_as_df_gdxdump('p_groupPolicy')
    index = 0
    for i in range(0,len(my_groups.index)):
        offset = 2
        while(True):
            if(my_weights.iloc[i,0] == p_groupPolicy['A{}'.format(offset)].value):
                break
            if(p_groupPolicy['A{}'.format(offset)].value == 'None'):
                print("ERROR: End of excel data")
                return
            offset += 1
            if(offset > 100):
                print("ERROR: no excel value found!")
                return
        p_groupPolicy['A{}'.format(offset)] = my_weights.iloc[i,0]
        p_groupPolicy['B{}'.format(offset)] = my_weights.iloc[i,1]
        p_groupPolicy['C{}'.format(offset)] = my_weights.iloc[i,2]

    input_data_file.update_gams_parameter_in_db("p_groupPolicy")
    shutil.copy2(input_data_file, os.path.join(input_dir, output_file))
    # target_gdx = os.path.join(input_folder_path, 'inputData_MGA.gdx')
    # input_data_file.save(os.path.join(input_dir, output_file))                        # save file under name containing emission cap
    return

#%%
#### Execute functions and definitions ####

## 0.1) Configuration
input_data = "inputData_MGA.gdx"   # input data file, save in input/costObj and input/MGA >>>>>>>>>> SAME NAME IN BOTH FOLDERS <<<<<<<<<<<
cost_slack = 10                    # in %, cost allowed in addition to minimum
number_alternatives = 10           # number of alternative solutions computed
year = 2030

## choose MGA method ##
# MGA_method = "normalised"        # Technology electricity variable divided by total electricity variable is weight for next interation
MGA_method = "integer"             # Adds +1 to each weight where the variable was > 0 in the previous run
# MGA_method = "min_max"           # Set every weight to -1/+1 at once while all other weights are 0
# MGA_method = "extrema"           # Set every weight to -1/+1 at once while all other weights are +1/-1                

variable = 'generation'            # choose whether to use capacity or generation decision variables, must be 'capacity' or 'generation'
                                   # capacity: use r_invest to determine Weights
                                   # generation: use r_gnuTotalGen to determine Weights
group_identifier = "weightGroup"   # all groups used for MGA must contain this identifying string

#%%
## 0.2) Some preparation
## gurantee that we have input data with valid groups, etc...
## 1) Run "normal" cost-minimising Backbone

# input_folder_path = r"C:\Users\oliver\Documents\RUB\01_Projekte\StEAM\Programme\StEAM_model\backbone-master\input\mga_input"
input_data = r'C:\Users\oliver\Documents\RUB\01_Projekte\StEAM\Programme\StEAM_model\backbone-master\input\mga_input\inputData_MGA.gdx'
input_data = 'inputData_MGA.gdx'
# output_folder_path = r"C:\Users\oliver\Documents\RUB\01_Projekte\StEAM\Programme\StEAM_model\backbone-master\input\mga_output"
output_data = r'C:\Users\oliver\Documents\RUB\01_Projekte\StEAM\Programme\StEAM_model\backbone-master\input\mga_input\result_costObj.gdx'
output_data = "result_costObj.gdx"

print(str(input_folder_path))
print(str(input_data))
print(str(output_folder_path))
print(str(output_data))
result = runBackbone(input_folder_path, input_data, output_folder_path, output_data)
print(result.stdout.decode('utf-8'))

if 'costMinResult' not in locals() or 'costMinResult' not in globals():
    # Check if the GDX file exists
    if os.path.exists(os.path.join(output_folder_path, output_data)):
        costMinResult = BackboneResult(os.path.join(output_folder_path, output_data))
        print(f"Reading result from GDX file: {os.path.join(output_folder_path, output_data)}")
    else:
        raise FileNotFoundError(f"The GDX file {os.path.join(output_folder_path, output_data)} does not exist.")
else:
    print("The 'result' variable is already defined.")

#%%
## 2) Read from results and determine max cost allowed
# costMinResult = BackboneResult(f"{output_folder_path}result_costObj.gdx")

# read groups from input data file, based on group_identifier
# uGroup = pd.read_excel(f"{input_folder_path}{input_data}", sheet_name="uGroup") 
ws = GamsWorkspace()
db = ws.add_database_from_gdx(os.path.join(input_folder_path, input_data))
d = list( (tuple(rec.keys)) for rec in db["uGroup"] )
uGroup = pd.DataFrame({'unit':[x for x,y in d], 'group':[y for x,y in d]})
groups = pd.DataFrame()
weight_indices = []
for i in uGroup.index:
    try:    # break if empty row (==NaN)
        if uGroup.iloc[i, 1].__contains__(group_identifier):
            weight_indices.append(i)
    except AttributeError:
        break
    except:
        print("Error caused by reading uGroup in inputData in /MGA folder")  # Use forward slashes
not_weight_indices = [i for i in uGroup.index if i not in weight_indices]
groups = pd.concat([groups,uGroup.iloc[weight_indices,1]], ignore_index=True)
groups.drop_duplicates(inplace=True)
uGroup.drop(index=not_weight_indices, inplace=True)
uGroup.set_index("group", inplace=True)

#%%
# initially set up weights dataframe to later write to input data
weights = pd.DataFrame(index=None, columns=["group", "param_policy", "value"])

weights["group"] = groups.loc[:,0]

if variable == 'capacity':
    weights["param_policy"] = "capacityWeight_MGA"
elif variable == 'generation':
    weights["param_policy"] = "generationWeight_MGA"
else:
    print("variable must be either 'capacity' or 'generation'")
weights["value"] = 0.
weights.drop_duplicates(inplace=True)
weights.set_index(np.arange(0,len(weights.index)),inplace=True)

minCost = round(costMinResult.param_as_df("r_cost_realizedCost").loc[0, "Val"], 3) # read parameter from results file
print("cost minimum (MEUR):", minCost)
maxCost = ((100 + cost_slack) / 100) * minCost
print(f"max cost allowed, including {cost_slack}% slack (MEUR):", maxCost)

cap = costMinResult.param_as_df("r_invest_unitCapacity_gnu")
cap.set_index("unit", inplace=True)

gen = costMinResult.param_as_df("r_gen_gnu")
gen.set_index("unit", inplace=True)
gen = gen[gen.Val >= 0]

#%%
## Loop over steps 3) to 5) KATHI

def add_values(variable, identifier):
    if(variable == "capacity"):
        try:
            iterator = iter(cap.loc[identifier,"Val"])
        except TypeError:
            return cap.loc[identifier,"Val"]
        except KeyError:
            return 0.
        else:
            return sum(cap.loc[identifier,"Val"])
    elif variable == "generation":
        try:
            iterator = iter(gen.loc[identifier,"Val"])
        except TypeError:
            return gen.loc[identifier,"Val"]
        except KeyError:
            return 0.
        else:
            return sum(gen.loc[identifier,"Val"])
                

if (MGA_method == "normalised") or (MGA_method == "integer"):
    iteration = 0
    while iteration < number_alternatives:    
        iteration += 1
        sum_weights = 0
        weights_tmp = dict()    
        # 3) Add capacity or generation from previous run to the weights and write new input data file
        if MGA_method == "normalised":
            for i in weights.index:
                g = weights.iloc[i,0]
                value = 0
                if isinstance(uGroup.loc[g, "unit"], str):
                    u = uGroup.loc[g, "unit"]
                    value += add_values(variable, u)
                else:
                    for u in uGroup.loc[g,"unit"]:
                        value += add_values(variable,u)
                weights_tmp[i] = value
                sum_weights += value
            for i in weights.index:
                weights.loc[i,"value"] += weights_tmp[i]/sum_weights

        if MGA_method == "integer":
            int_tol = 0.001
            for i in weights.index:
                g = weights.iloc[i,0]
                value = 0
                if isinstance(uGroup.loc[g, "unit"], str):
                    u = uGroup.loc[g, "unit"]
                    value += add_values(variable, u)
                else:
                    for u in uGroup.loc[g,"unit"]:
                        value += add_values(variable,u)
                weights_tmp[i] = float((abs(value) > int_tol))
            for i in weights.index:
                weights.loc[i,"value"] += weights_tmp[i]

        print("Weights for iteration", iteration)
        print(weights)

        writeWeightsToInputDataInBetter(input_folder_path, input_data, f"{input_data}_it{iteration}.gdx", weights, groups)

        # 4) Run Backbone with "new" objective function
        # objective function is weighted sum of decision variables (capacity or generation)
        runBackbone(input_folder_path, f"{input_data}_it{iteration}.gdx", output_folder_path, f"result_it{iteration}.gdx",
                    costConstraint=True, maxCost=str(maxCost))
        # 5) Read from latest results
        bb_result = BackboneResult(f"{output_folder_path}result_it{iteration}.gdx")
        cap = bb_result.param_as_df("r_invest")
        cap.set_index("unit", inplace=True)
        gen = bb_result.param_as_df("r_gnuTotalGen")
        gen.set_index("unit", inplace=True)
        gen = gen[gen.Val >= 0]
    
if MGA_method == "min_max":
    min_max_weights = [-1, 1]
    iteration = 1
    for i in weights.index:
        for w in min_max_weights:
            weights.iloc[i,2] = w
            writeWeightsToInputDataInBetter(input_folder_path, input_data, f"{input_data}_it{iteration}.gdx", weights, groups)

            # 4) Run Backbone with "new" objective function
            runBackbone(input_folder_path, f"{input_data}_it{iteration}.gdx", input_folder_path, f"result_it{iteration}.gdx",
                        costConstraint=True, maxCost=str(maxCost))
            iteration +=1 
            print("index of weight", i, "min_max weight", w)
            if w == 1:
                weights.iloc[i,2] = 0
                break


if MGA_method == "extrema":
    weights["value"] = 1
    iteration = 1
    for i in weights.index:
        weights.iloc[i,2] = -1
        writeWeightsToInputDataInBetter(input_folder_path, input_data, f"{input_data}_it{iteration}.gdx", weights, groups)

        # 4) Run Backbone with "new" objective function
        runBackbone(input_folder_path, f"{input_data}_it{iteration}.gdx", output_folder_path, f"result_it{iteration}.gdx",
                    costConstraint=True, maxCost=str(maxCost))
        iteration +=1
        weights.iloc[i,2] = 1 
        print("index of weight", i)
    weights["value"] = -1
    for i in weights.index:
        weights.iloc[i,2] = 1
        writeWeightsToInputDataInBetter(input_folder_path, input_data, f"{input_data}_it{iteration}.gdx", weights, groups)

        # 4) Run Backbone with "new" objective function
        runBackbone(input_folder_path, f"{input_data}_it{iteration}.gdx", output_folder_path, f"result_it{iteration}.gdx",
                    costConstraint=True, maxCost=str(maxCost))
        iteration +=1
        weights.iloc[i,2] = -1 
        print("index of weight", i)

#%%